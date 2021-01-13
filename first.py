from PIL import Image
import numpy as np
import pytesseract
import re
import argparse
import cv2
import os
import openpyxl
import xlrd
from xlutils.copy import copy
from datetime import datetime
from datetime import date
import  imutils


wb=openpyxl.load_workbook("WorkBook.xlsx")
s=wb.worksheets[0]
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
today = date.today()
current_date = today.strftime("%d/%m/%Y")


# Read the image file
image = cv2.imread("Path to vehicle's image")

# Resize the image - change width to 500
image = imutils.resize(image, width=500)

# Display the original image
cv2.imshow("Original Image", image)

# RGB to Gray scale conversion
gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
cv2.imshow("1 - Grayscale Conversion", gray)

# Noise removal with iterative bilateral filter(removes noise while preserving edges)
gray = cv2.bilateralFilter(gray, 11, 17, 17)
cv2.imshow("2 - Bilateral Filter", gray)

# Noise removal with iterative bilateral filter(removes noise while preserving edges)
noise_removal = cv2.bilateralFilter(gray,9,75,75)
cv2.namedWindow("2 - Noise Removal(Bilateral Filtering)",cv2.WINDOW_NORMAL)
cv2.imshow("2 - Noise Removal(Bilateral Filtering)",noise_removal)

# Histogram equalisation for better results
equal_histogram = cv2.equalizeHist(noise_removal)
cv2.namedWindow("3 - Histogram equalisation",cv2.WINDOW_NORMAL)
cv2.imshow("3 - Histogram equalisation",equal_histogram)

# Morphological opening with a rectangular structure element
kernel = cv2.getStructuringElement(cv2.MORPH_RECT,(5,5))                                # create the kernel
morph_image = cv2.morphologyEx(equal_histogram,cv2.MORPH_OPEN,kernel,iterations=15)     # Morphological opening using the kernal created
cv2.namedWindow("4 - Morphological opening",cv2.WINDOW_NORMAL)
cv2.imshow("4 - Morphological opening",morph_image)

# Image subtraction(Subtracting the Morphed image from the histogram equalised Image)
sub_morp_image = cv2.subtract(equal_histogram,morph_image)
cv2.namedWindow("5 - Image Subtraction", cv2.WINDOW_NORMAL)
cv2.imshow("5 - Image Subtraction", sub_morp_image)

# Thresholding the image
ret,thresh_image = cv2.threshold(sub_morp_image,0,255,cv2.THRESH_OTSU)
cv2.namedWindow("6 - Thresholding",cv2.WINDOW_NORMAL)
cv2.imshow("6 - Thresholding",thresh_image)

# Find Edges of the grayscale image
edged = cv2.Canny(gray, 170, 200)
cv2.imshow("4 - Canny Edges", edged)

# Find contours based on Edges
cnts, heirarchy = cv2.findContours(edged.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
cnts=sorted(cnts, key = cv2.contourArea, reverse = True)[:30] #sort contours based on their area keeping minimum required area as '30' (anything smaller than this will not be considered)
NumberPlateCnt = 0 #we currently have no Number plate contour

# loop over our contours to find the best possible approximate contour of number plate
count = 0
for c in cnts:
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.02 * peri, True)
        if len(approx) == 4:  # Select the contour with 4 corners
            NumberPlateCnt = approx #This is our approx Number Plate Contour
            break


# Drawing the selected contour on the original image
cv2.drawContours(image, [NumberPlateCnt], -1, (0,255,0), 3)
cv2.imshow("Final Image With Number Plate Detected", image)


# SEPARATING OUT THE NUMBER PLATE FROM IMAGE:

# Masking the part other than the number plate
mask = np.zeros(gray.shape,np.uint8)                            # create an empty black image
new_image = cv2.drawContours(mask,[NumberPlateCnt],0,255,-1,)       # Draw the contour of number plate on the black image - This is our mask
new_image = cv2.bitwise_and(image,image,mask=mask)                      # Take bitwise AND with the original image so we can just get the Number Plate from the original image
cv2.namedWindow("10 - Number Plate Separation",cv2.WINDOW_NORMAL)
cv2.imshow("10 - Number Plate Separation",new_image)



#HISTOGRAM EQUALIZATION FOR ENHANCING THE NUMBER PLATE FOR FURTHER PROCESSING:


y,cr,cb = cv2.split(cv2.cvtColor(new_image,cv2.COLOR_RGB2YCrCb))        # Converting the image to YCrCb model and splitting the 3 channels
y = cv2.equalizeHist(y)                                                 # Applying histogram equalisation
final_image = cv2.cvtColor(cv2.merge([y,cr,cb]),cv2.COLOR_YCrCb2RGB)    # Merging the 3 channels
cv2.namedWindow("11 - Enhanced Number Plate",cv2.WINDOW_NORMAL)
cv2.imshow("11 - Enhanced Number Plate",image)



pytesseract.pytesseract.tesseract_cmd = 'path to tesseract installion'
# write the grayscale image to disk as a temporary file so we can
# apply OCR to it
filename = "{}.png".format(os.getpid())
cv2.imwrite(filename,final_image)
text = pytesseract.image_to_string(Image.open(filename),config='--psm 6 --oem 3 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ')
text_strip = ""
for e in text :
    if(len(text_strip)> 10) :
       break;
    if(e.isalnum()) :
        text_strip+=e
#text_strip = text_strip[1:]
print(text_strip.upper())
#cv2.imshow("Image", image)
#cv2.imshow("Output", gray)
cv2.waitKey(0)
for columns in range(2) :
            for rows in range (100) :
                if(s.cell(row=rows+1,column=columns+1).value==None and s.cell(row=rows+1,column=columns+2).value==None) :
                    break
            break
                    
s.cell(row=rows+1,column=columns+1).value= text_strip
s.cell(row=rows+1,column=columns+2).value= current_time
s.cell(row=rows+1,column=columns+3).value= current_date
wb.save("EDI_template.xlsx")
                   
